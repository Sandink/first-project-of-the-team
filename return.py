import flask
from flask import request, jsonify, make_response
from func import main as mn
from func import ab
from flask_cors import *
import table_xy
from openpyxl import Workbook,load_workbook

server = flask.Flask('__name__')
CORS(server, supports_credentials=True)

@server.route('/login/file',methods=['get','post'])
def file():
    files = request.files.get(r'file')

    # 加载文件
    wb = load_workbook(files)
    # 激活默认工作表
    ws = wb.active
    # 获取工作表的行列数
    x = ws.max_row
    y = ws.max_column

    # 定义新工作表的位置
    path = r'./函数数据.xlsx'
    # 创建新的工作簿
    wb_new = Workbook()
    # 选择默认工作表激活
    ws_new = wb_new.active

    # 对工作进行遍历，读取所有数字
    for i in range(1,x+1):
        for j in range(1,y+1):
            ws_new.cell(i,j,ws.cell(i,j).value)

    wb_new.save(path)
    wb_new.close()

    s = 'over!!!!'
    response = make_response(jsonify(s))
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = '*'
    response.headers['Access-Control-Allow-Headers'] = '*'
    return response

@server.route('/login/inputs',methods=['get','post'])
def inputs():
    # 获取通过url请求传参的数据
    x = request.values.get('xvalue')
    try:
        xx = float(x)
        resu = mn(xx)
        response = make_response(jsonify(resu))
    except Exception as e:
        # resu = "'您输入的数据类型不合法，我无法为您返回数据！':-1"
        # 将字典转换为json串, json是字符串
        # return jsonify(-1)
        response = make_response(jsonify(-1))

    # response = make_response(jsonify([x,y]))
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'OPTIONS,HEAD,GET,POST'
    response.headers['Access-Control-Allow-Headers'] = 'x-requested-with'
    return response

@server.route('/login/array',methods=['get','post'])
def arrays():
    x,y = table_xy.table_xy()
    # return jsonify([x,y])

    response = make_response(jsonify([x,y]))
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'OPTIONS,HEAD,GET,POST'
    response.headers['Access-Control-Allow-Headers'] = 'x-requested-with'
    return response

@server.route('/login/newArray',methods=['get','post'])
def newArrays():
    x,y = table_xy.table_xy()
    a,b = ab()
    new_y = []
    for i in x:
        # ny = b * i + a
        ny = float('{:.3f}'.format(b * i + a))
        new_y.append(ny)

    response = make_response(jsonify([x,new_y]))
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'OPTIONS,HEAD,GET,POST'
    response.headers['Access-Control-Allow-Headers'] = 'x-requested-with'
    return response
@server.route('/login/get_ab',methods=['get','post'])
def canshu():
    a,b = ab()

    response = make_response(jsonify(a,b))
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'OPTIONS,HEAD,GET,POST'
    response.headers['Access-Control-Allow-Headers'] = 'x-requested-with'
    return response

if __name__ == '__main__':
    server.run(debug=True,port=8223,host='0.0.0.0')