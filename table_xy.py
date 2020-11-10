from openpyxl import Workbook,load_workbook

def table_xy():

    # 加载文件
    wb = load_workbook(r'./函数数据.xlsx')
    # 激活默认工作表
    ws = wb.active
    # 获取工作表的行数
    x = ws.max_row

    # 定义两个列表，分别存储x值与y值
    xvalue = []
    yvalue = []

    for i in range(2,x+1):
        xvalue.append(ws.cell(i,1).value)
        yvalue.append(ws.cell(i,2).value)

    # print(xvalue,yvalue)
    return xvalue,yvalue